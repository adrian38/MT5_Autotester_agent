from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import math
import re
import statistics

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .mt5_report import StrategyReport, Trade


PORTFOLIO_ACCOUNT_BALANCE = 1000.0

SUMMARY_HEADERS = [
    "Symbol",
    "Timeframe",
    "Period start",
    "Period end",
    "Initial deposit",
    "Total profit",
    "# of trades",
    "Sharpe ratio",
    "Profit factor",
    "Return / DD ratio",
    "Winning %",
    "Profit in pips",
    "Yearly avg profit",
    "Drawdown",
    "Percent drawdown",
    "Daily avg profit",
    "Monthly avg profit",
    "Average trade",
    "Yearly avg % return",
    "CAGR",
]


@dataclass
class DrawdownWindow:
    trades: list[Trade]
    peak_balance: float
    trough_balance: float
    drawdown: float
    start_time: datetime | None
    end_time: datetime | None


@dataclass
class StrategyDayDrawdown:
    report: StrategyReport
    window: DrawdownWindow


@dataclass
class PortfolioDayDrawdown:
    day: object
    total_drawdown: float
    total_profit: float
    strategies: list[StrategyDayDrawdown]


@dataclass
class PortfolioValleyContribution:
    report: StrategyReport
    trades: list[Trade]
    total_profit: float


@dataclass
class PortfolioValleyDrawdown:
    peak_time: datetime | None
    trough_time: datetime | None
    initial_balance: float
    peak_balance: float
    trough_balance: float
    drawdown: float
    total_profit: float
    contributions: list[PortfolioValleyContribution]


def build_drawdown_workbook(reports: list[StrategyReport], output_path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    used_names: set[str] = set()
    for report in reports:
        ws = wb.create_sheet(_safe_sheet_name(report.name, used_names))
        window = max_drawdown_window(report)
        _write_dd_sheet(ws, report, window)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_portfolio_drawdown_workbook(reports: list[StrategyReport], output_path: Path) -> None:
    portfolio_dd = max_portfolio_drawdown_day(reports)
    wb = Workbook()
    ws = wb.active
    ws.title = "PORTFOLIO_DD"
    _write_portfolio_dd_sheet(ws, portfolio_dd)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_portfolio_valley_drawdown_workbook(reports: list[StrategyReport], output_path: Path) -> None:
    portfolio_dd = max_portfolio_valley_drawdown(reports)
    wb = Workbook()
    ws = wb.active
    ws.title = "PORTFOLIO_VALLEY_DD"
    _write_portfolio_valley_dd_sheet(ws, portfolio_dd)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_top_portfolio_valleys_workbook(reports: list[StrategyReport], output_path: Path, limit: int = 5) -> None:
    valleys = top_portfolio_valley_drawdowns(reports, limit)
    wb = Workbook()
    ws = wb.active
    ws.title = "TOP_VALLEYS"
    _write_top_portfolio_valleys_sheet(ws, valleys)
    for index, valley in enumerate(valleys, start=1):
        detail = wb.create_sheet(f"VALLE_{index}")
        _write_portfolio_valley_dd_sheet(detail, valley)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_dd_threshold_workbook(reports: list[StrategyReport], output_path: Path, threshold: float) -> None:
    rows = []
    for report in reports:
        window = max_drawdown_window(report)
        stats = _window_stats(report, window)
        rows.append(
            {
                "Strategy": report.name,
                "Symbol": report.symbol,
                "Timeframe": report.timeframe,
                "Cumple": "SI" if window.drawdown <= threshold else "NO",
                "Threshold": threshold,
                "Max daily DD": round(window.drawdown, 2),
                "Worst day P/L": stats["Total profit"],
                "Worst day": stats["Period start"],
                "# trades": stats["# of trades"],
                "Winning %": stats["Winning %"],
                "Profit factor": stats["Profit factor"],
                "Profit in pips": stats["Profit in pips"],
                "Report path": str(report.path),
            }
        )

    rows.sort(key=lambda row: (row["Cumple"] != "SI", row["Max daily DD"], row["Strategy"]))

    wb = Workbook()
    ws_ok = wb.active
    ws_ok.title = "CUMPLEN"
    _write_threshold_sheet(ws_ok, [row for row in rows if row["Cumple"] == "SI"], threshold)
    ws_all = wb.create_sheet("TODAS")
    _write_threshold_sheet(ws_all, rows, threshold)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def max_portfolio_drawdown_day(reports: list[StrategyReport]) -> PortfolioDayDrawdown:
    by_day: dict[object, list[StrategyDayDrawdown]] = {}

    for report in reports:
        for day, window in daily_close_windows(report).items():
            by_day.setdefault(day, []).append(StrategyDayDrawdown(report, window))

    if not by_day:
        return PortfolioDayDrawdown("", 0.0, 0.0, [])

    best_day = None
    worst_total_profit = 0.0
    for day, strategy_windows in by_day.items():
        total_profit = sum(sum(trade.profit_loss for trade in item.window.trades) for item in strategy_windows)
        if best_day is None or total_profit < worst_total_profit:
            best_day = day
            worst_total_profit = total_profit

    assert best_day is not None
    strategies = sorted(
        by_day[best_day],
        key=lambda item: sum(trade.profit_loss for trade in item.window.trades),
    )
    total_profit = sum(sum(trade.profit_loss for trade in item.window.trades) for item in strategies)
    return PortfolioDayDrawdown(best_day, max(-total_profit, 0.0), total_profit, strategies)


def max_portfolio_valley_drawdown(reports: list[StrategyReport]) -> PortfolioValleyDrawdown:
    valleys = top_portfolio_valley_drawdowns(reports, 1)
    if valleys:
        return valleys[0]

    return PortfolioValleyDrawdown(None, None, PORTFOLIO_ACCOUNT_BALANCE, PORTFOLIO_ACCOUNT_BALANCE, PORTFOLIO_ACCOUNT_BALANCE, 0.0, 0.0, [])


def top_portfolio_valley_drawdowns(reports: list[StrategyReport], limit: int = 5) -> list[PortfolioValleyDrawdown]:
    portfolio_trades: list[tuple[StrategyReport, Trade]] = []
    for report in reports:
        for trade in report.trades:
            portfolio_trades.append((report, trade))
    portfolio_trades.sort(key=lambda item: item[1].close_time)

    if not portfolio_trades:
        return []

    balance = PORTFOLIO_ACCOUNT_BALANCE
    peak_balance = PORTFOLIO_ACCOUNT_BALANCE
    peak_time: datetime | None = None
    peak_index = -1
    trough_balance = PORTFOLIO_ACCOUNT_BALANCE
    trough_time: datetime | None = None
    trough_index = -1
    valleys: list[PortfolioValleyDrawdown] = []

    for index, (_report, trade) in enumerate(portfolio_trades):
        balance += trade.profit_loss
        if balance > peak_balance:
            if trough_index > peak_index and peak_balance > trough_balance:
                valleys.append(
                    _portfolio_valley_from_indices(
                        portfolio_trades,
                        peak_index,
                        trough_index,
                        peak_time,
                        trough_time,
                        peak_balance,
                        trough_balance,
                    )
                )
            peak_balance = balance
            peak_time = trade.close_time
            peak_index = index
            trough_balance = balance
            trough_time = trade.close_time
            trough_index = index
        elif balance < trough_balance:
            trough_balance = balance
            trough_time = trade.close_time
            trough_index = index

    if trough_index > peak_index and peak_balance > trough_balance:
        valleys.append(
            _portfolio_valley_from_indices(
                portfolio_trades,
                peak_index,
                trough_index,
                peak_time,
                trough_time,
                peak_balance,
                trough_balance,
            )
        )

    valleys.sort(key=lambda valley: valley.drawdown, reverse=True)
    return valleys[:limit]


def _portfolio_valley_from_indices(
    portfolio_trades: list[tuple[StrategyReport, Trade]],
    peak_index: int,
    trough_index: int,
    peak_time: datetime | None,
    trough_time: datetime | None,
    peak_balance: float,
    trough_balance: float,
) -> PortfolioValleyDrawdown:
    selected = portfolio_trades[peak_index + 1 : trough_index + 1] if trough_index >= 0 else []
    grouped: dict[Path, tuple[StrategyReport, list[Trade]]] = {}
    for report, trade in selected:
        if report.path not in grouped:
            grouped[report.path] = (report, [])
        grouped[report.path][1].append(trade)

    contributions = [
        PortfolioValleyContribution(report, trades, sum(trade.profit_loss for trade in trades))
        for report, trades in grouped.values()
    ]
    contributions.sort(key=lambda item: item.total_profit)

    return PortfolioValleyDrawdown(
        peak_time=peak_time,
        trough_time=trough_time,
        initial_balance=PORTFOLIO_ACCOUNT_BALANCE,
        peak_balance=peak_balance,
        trough_balance=trough_balance,
        drawdown=peak_balance - trough_balance,
        total_profit=sum(item.total_profit for item in contributions),
        contributions=contributions,
    )


def _write_threshold_sheet(ws, rows: list[dict[str, object]], threshold: float) -> None:
    ws["A1"] = "Filtro por pico diario maximo"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Umbral DD diario"
    ws["B3"] = threshold
    ws["A4"] = "Estrategias"
    ws["B4"] = len(rows)
    ws["A3"].font = Font(bold=True)
    ws["A4"].font = Font(bold=True)

    headers = [
        "Strategy",
        "Symbol",
        "Timeframe",
        "Cumple",
        "Threshold",
        "Max daily DD",
        "Worst day P/L",
        "Worst day",
        "# trades",
        "Winning %",
        "Profit factor",
        "Profit in pips",
        "Report path",
    ]
    header_row = 7
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row_idx, col, row[header])
            cell.border = _thin_border()
            if header == "Cumple":
                cell.fill = PatternFill("solid", fgColor="C6EFCE" if row[header] == "SI" else "FFC7CE")
            if header in {"Threshold", "Max daily DD", "Worst day P/L"}:
                cell.number_format = '$ #,##0.00;[Red]$ -#,##0.00;$ 0.00'

    ws.freeze_panes = "A8"
    if rows:
        ws.auto_filter.ref = f"A{header_row}:M{header_row + len(rows)}"
    _set_widths(ws, [30, 12, 12, 10, 12, 14, 14, 13, 10, 12, 12, 13, 72])


def max_drawdown_window(report: StrategyReport) -> DrawdownWindow:
    windows = daily_close_windows(report)
    if not windows:
        return DrawdownWindow([], report.initial_deposit, report.initial_deposit, 0.0, None, None)

    return max(windows.values(), key=lambda window: window.drawdown)


def daily_drawdown_windows(report: StrategyReport) -> dict[object, DrawdownWindow]:
    trades = sorted(report.trades, key=lambda trade: trade.close_time)
    if not trades:
        return {}

    balance = report.initial_deposit
    windows: dict[object, DrawdownWindow] = {}
    day_trades: list[tuple[Trade, float, float]] = []
    current_day = trades[0].close_time.date()

    for trade in trades:
        trade_day = trade.close_time.date()
        if trade_day != current_day:
            candidate = _daily_drawdown_window(day_trades)
            if candidate:
                windows[current_day] = candidate
            day_trades = []
            current_day = trade_day

        before = balance
        balance += trade.profit_loss
        day_trades.append((trade, before, balance))

    candidate = _daily_drawdown_window(day_trades)
    if candidate:
        windows[current_day] = candidate
    return windows


def daily_close_windows(report: StrategyReport) -> dict[object, DrawdownWindow]:
    trades = sorted(report.trades, key=lambda trade: trade.close_time)
    if not trades:
        return {}

    windows: dict[object, DrawdownWindow] = {}
    grouped: dict[object, list[Trade]] = {}
    for trade in trades:
        grouped.setdefault(trade.close_time.date(), []).append(trade)

    for day, day_trades in grouped.items():
        total_profit = sum(trade.profit_loss for trade in day_trades)
        close_time_start = day_trades[0].close_time
        close_time_end = day_trades[-1].close_time
        windows[day] = DrawdownWindow(
            trades=day_trades,
            peak_balance=report.initial_deposit,
            trough_balance=report.initial_deposit + total_profit,
            drawdown=max(-total_profit, 0.0),
            start_time=close_time_start,
            end_time=close_time_end,
        )
    return windows


def _write_portfolio_dd_sheet(ws, portfolio_dd: PortfolioDayDrawdown) -> None:
    day_text = portfolio_dd.day.strftime("%d.%m.%Y") if hasattr(portfolio_dd.day, "strftime") else str(portfolio_dd.day)
    ws["A1"] = "Portfolio DD puntual"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Dia max DD"
    ws["B3"] = day_text
    ws["A4"] = "Portfolio DD"
    ws["B4"] = round(portfolio_dd.total_drawdown, 2)
    ws["A5"] = "P/L neto del dia"
    ws["B5"] = round(portfolio_dd.total_profit, 2)
    ws["A6"] = "# estrategias con cierres"
    ws["B6"] = len(portfolio_dd.strategies)

    for cell_ref in ["A3", "A4", "A5", "A6"]:
        ws[cell_ref].font = Font(bold=True)
    for cell_ref in ["B4", "B5"]:
        ws[cell_ref].number_format = '$ #,##0.00;[Red]$ -#,##0.00;$ 0.00'

    headers = [
        "Strategy",
        "Symbol",
        "Timeframe",
        "Period start",
        "Period end",
        "DD contribution",
        "Total profit",
        "# trades",
        "Winning %",
        "Profit factor",
        "Profit in pips",
        "Average trade",
        "Report path",
    ]
    header_row = 9
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row, item in enumerate(portfolio_dd.strategies, start=header_row + 1):
        report = item.report
        window = item.window
        stats = _window_stats(report, window)
        total_profit = sum(trade.profit_loss for trade in window.trades)
        dd_contribution = -total_profit
        values = [
            report.name,
            report.symbol,
            report.timeframe,
            stats["Period start"],
            stats["Period end"],
            round(dd_contribution, 2),
            round(total_profit, 2),
            stats["# of trades"],
            stats["Winning %"],
            stats["Profit factor"],
            stats["Profit in pips"],
            stats["Average trade"],
            str(report.path),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = _thin_border()
            if col in {6, 7, 12}:
                cell.number_format = '$ #,##0.00;[Red]$ -#,##0.00;$ 0.00'

    total_row = header_row + 1 + len(portfolio_dd.strategies)
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(total_row, 6, round(portfolio_dd.total_drawdown, 2)).font = Font(bold=True)
    ws.cell(total_row, 7, round(portfolio_dd.total_profit, 2)).font = Font(bold=True)
    ws.freeze_panes = "A10"
    ws.auto_filter.ref = f"A{header_row}:M{max(header_row, total_row - 1)}"
    _set_widths(ws, [30, 12, 12, 13, 13, 13, 13, 10, 12, 12, 13, 13, 72])


def _write_portfolio_valley_dd_sheet(ws, portfolio_dd: PortfolioValleyDrawdown) -> None:
    ws["A1"] = "Portfolio DD por picos"
    ws["A1"].font = Font(bold=True, size=16)
    summary = [
        ("Pico", _datetime_text(portfolio_dd.peak_time)),
        ("Valle", _datetime_text(portfolio_dd.trough_time)),
        ("Cuenta test", round(portfolio_dd.initial_balance, 2)),
        ("Neto cerrado pico", round(portfolio_dd.peak_balance - portfolio_dd.initial_balance, 2)),
        ("Neto cerrado valle", round(portfolio_dd.trough_balance - portfolio_dd.initial_balance, 2)),
        ("Balance pico", round(portfolio_dd.peak_balance, 2)),
        ("Balance valle", round(portfolio_dd.trough_balance, 2)),
        ("DD valle", round(portfolio_dd.drawdown, 2)),
        ("P/L neto valle", round(portfolio_dd.total_profit, 2)),
        ("# estrategias afectadas", len(portfolio_dd.contributions)),
    ]
    for row, (label, value) in enumerate(summary, start=3):
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, value)
    for row in [5, 6, 7, 8, 9, 10]:
        ws.cell(row, 2).number_format = '$ #,##0.00;[Red]$ -#,##0.00;$ 0.00'

    headers = [
        "Strategy",
        "Symbol",
        "Timeframe",
        "Period start",
        "Period end",
        "DD contribution",
        "Total profit",
        "# trades",
        "Winning %",
        "Profit factor",
        "Profit in pips",
        "Average trade",
        "Report path",
    ]
    header_row = 14
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row, item in enumerate(portfolio_dd.contributions, start=header_row + 1):
        report = item.report
        trades = sorted(item.trades, key=lambda trade: trade.close_time)
        window = DrawdownWindow(
            trades=trades,
            peak_balance=portfolio_dd.peak_balance,
            trough_balance=portfolio_dd.trough_balance,
            drawdown=-item.total_profit,
            start_time=trades[0].close_time if trades else None,
            end_time=trades[-1].close_time if trades else None,
        )
        stats = _window_stats(report, window)
        values = [
            report.name,
            report.symbol,
            report.timeframe,
            stats["Period start"],
            stats["Period end"],
            round(-item.total_profit, 2),
            round(item.total_profit, 2),
            stats["# of trades"],
            stats["Winning %"],
            stats["Profit factor"],
            stats["Profit in pips"],
            stats["Average trade"],
            str(report.path),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = _thin_border()
            if col in {6, 7, 12}:
                cell.number_format = '$ #,##0.00;[Red]$ -#,##0.00;$ 0.00'

    total_row = header_row + 1 + len(portfolio_dd.contributions)
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(total_row, 6, round(portfolio_dd.drawdown, 2)).font = Font(bold=True)
    ws.cell(total_row, 7, round(portfolio_dd.total_profit, 2)).font = Font(bold=True)
    for col in [6, 7]:
        ws.cell(total_row, col).number_format = '$ #,##0.00;[Red]$ -#,##0.00;$ 0.00'
    ws.freeze_panes = "A15"
    ws.auto_filter.ref = f"A{header_row}:M{max(header_row, total_row - 1)}"
    _set_widths(ws, [30, 12, 12, 13, 13, 14, 13, 10, 12, 12, 13, 13, 72])


def _write_top_portfolio_valleys_sheet(ws, valleys: list[PortfolioValleyDrawdown]) -> None:
    ws["A1"] = "Top 5 peores valles del portfolio"
    ws["A1"].font = Font(bold=True, size=16)
    headers = [
        "Rank",
        "Pico",
        "Valle",
        "Neto cerrado pico",
        "Neto cerrado valle",
        "Balance pico",
        "Balance valle",
        "DD valle",
        "P/L neto valle",
        "# estrategias afectadas",
    ]
    header_row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row, valley in enumerate(valleys, start=header_row + 1):
        values = [
            row - header_row,
            _datetime_text(valley.peak_time),
            _datetime_text(valley.trough_time),
            round(valley.peak_balance - valley.initial_balance, 2),
            round(valley.trough_balance - valley.initial_balance, 2),
            round(valley.peak_balance, 2),
            round(valley.trough_balance, 2),
            round(valley.drawdown, 2),
            round(valley.total_profit, 2),
            len(valley.contributions),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = _thin_border()
            if col in {4, 5, 6, 7, 8, 9}:
                cell.number_format = '$ #,##0.00;[Red]$ -#,##0.00;$ 0.00'

    ws.freeze_panes = "A4"
    if valleys:
        ws.auto_filter.ref = f"A{header_row}:J{header_row + len(valleys)}"
    _set_widths(ws, [8, 20, 20, 15, 15, 15, 15, 15, 15, 20])



def _daily_drawdown_window(day_trades: list[tuple[Trade, float, float]]) -> DrawdownWindow | None:
    if not day_trades:
        return None

    day_open_balance = day_trades[0][1]
    trough_index = 0
    trough_balance = day_trades[0][2]
    for index, (_trade, _before, after) in enumerate(day_trades):
        if after < trough_balance:
            trough_index = index
            trough_balance = after

    dd = max(day_open_balance - trough_balance, 0.0)
    if dd <= 0:
        return DrawdownWindow([], day_open_balance, day_open_balance, 0.0, None, None)

    # Include every trade that closes on this DD day, even if it opened earlier.
    selected = [trade for trade, _before, _after in day_trades]
    start_time = selected[0].close_time if selected else None
    end_time = selected[-1].close_time if selected else None
    return DrawdownWindow(selected, day_open_balance, trough_balance, dd, start_time, end_time)


def _write_dd_sheet(ws, report: StrategyReport, window: DrawdownWindow) -> None:
    stats = _window_stats(report, window)
    ws["A1"] = f"Strategy: {report.name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Summary"
    ws["A3"].font = Font(bold=True)

    for col, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _thin_border()

    for col, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(5, col, stats[header])
        cell.border = _thin_border()

    units = ["", "", "date", "date", "$", "$", "", "", "", "", "%", "pips", "$", "$", "%", "$", "$", "$", "%", "%"]
    for col, unit in enumerate(units, start=1):
        ws.cell(6, col, unit)

    strategy_row = 8
    ws.cell(strategy_row, 1, "Stats - Strategy").font = Font(bold=True)
    strategy_headers = [
        "Wins/Losses Ratio",
        "Payout Ratio (Avg Win/Loss)",
        "Average # of Bars in Trade",
        "AHPR",
        "Z-Score",
        "Z-Probability",
        "Expectancy",
        "Deviation",
        "Exposure",
        "Stagnation in Days",
        "Stagnation in %",
    ]
    strategy_values = _strategy_stats(window.trades, report.initial_deposit)
    _write_horizontal_table(ws, strategy_row + 1, strategy_headers, strategy_values)

    trades_row = 13
    ws.cell(trades_row, 1, "Stats - Trades").font = Font(bold=True)
    trades_headers = [
        "# of Wins",
        "# of Losses",
        "# of Cancelled/Expired",
        "Gross Profit",
        "Gross Loss",
        "Average Win",
        "Average Loss",
        "Largest Win",
        "Largest Loss",
        "Max Consec Wins",
        "Max Consec Losses",
        "Avg Consec Wins",
        "Avg Consec Loss",
        "Avg # of Bars in Wins",
        "Avg # of Bars in Losses",
    ]
    trades_values = _trade_stats(window.trades)
    _write_horizontal_table(ws, trades_row + 1, trades_headers, trades_values)

    ws.freeze_panes = "A4"
    _set_widths(ws, [13, 14, 14, 14, 14, 13, 11, 13, 13, 15, 12, 13, 15, 13, 14, 13, 14, 13, 15, 10, 10, 10, 10, 10, 10])


def _window_stats(report: StrategyReport, window: DrawdownWindow) -> dict[str, float | int | str]:
    trades = window.trades
    profits = [trade.profit_loss for trade in trades]
    wins = [value for value in profits if value > 0]
    losses = [value for value in profits if value < 0]
    total_profit = round(sum(profits), 2)
    gross_profit = sum(wins)
    gross_loss = abs(sum(losses))
    start = window.start_time or (trades[0].open_time if trades else None)
    end = window.end_time or (trades[-1].close_time if trades else None)
    days = max(((end.date() - start.date()).days + 1), 1) if start and end else 1
    yearly_avg = total_profit / days * 365 if days else 0.0
    monthly_avg = total_profit / max(math.ceil(days / 30), 1)
    avg_trade = _avg(profits)
    deviation = statistics.pstdev(profits) if len(profits) > 1 else 0.0
    drawdown_pct = window.drawdown / window.peak_balance * 100 if window.peak_balance else 0.0
    yearly_pct = yearly_avg / report.initial_deposit * 100 if report.initial_deposit else 0.0

    return {
        "Symbol": report.symbol,
        "Timeframe": report.timeframe,
        "Period start": _date(start),
        "Period end": _date(end),
        "Initial deposit": round(report.initial_deposit, 2),
        "Total profit": total_profit,
        "# of trades": len(trades),
        "Sharpe ratio": round(avg_trade / deviation, 2) if deviation else round(avg_trade * 100, 2),
        "Profit factor": round(gross_profit / gross_loss, 2) if gross_loss else 0,
        "Return / DD ratio": round(total_profit / window.drawdown, 2) if window.drawdown else 0,
        "Winning %": round(len(wins) / len(profits) * 100, 2) if profits else 0,
        "Profit in pips": round(_profit_in_pips(report, trades), 1),
        "Yearly avg profit": round(yearly_avg, 2),
        "Drawdown": round(window.drawdown, 2),
        "Percent drawdown": round(drawdown_pct, 2),
        "Daily avg profit": round(total_profit / days, 2),
        "Monthly avg profit": round(monthly_avg, 2),
        "Average trade": round(avg_trade, 2),
        "Yearly avg % return": round(yearly_pct, 2),
        "CAGR": 0 if total_profit < 0 else round(yearly_pct, 2),
    }


def _strategy_stats(trades: list[Trade], initial_deposit: float) -> list[float | int]:
    profits = [trade.profit_loss for trade in trades]
    wins = [value for value in profits if value > 0]
    losses = [value for value in profits if value < 0]
    avg_trade = _avg(profits)
    deviation = statistics.pstdev(profits) if len(profits) > 1 else 0
    return [
        _safe_div(len(wins), len(losses)),
        _safe_div(_avg(wins), abs(_avg(losses))),
        0,
        round(avg_trade / initial_deposit * 100, 2) if initial_deposit else 0,
        0,
        99.9 if losses else 0,
        round(avg_trade, 2),
        round(deviation, 2),
        -999999999,
        0,
        0,
    ]


def _trade_stats(trades: list[Trade]) -> list[float | int]:
    profits = [trade.profit_loss for trade in trades]
    wins = [value for value in profits if value > 0]
    losses = [value for value in profits if value < 0]
    win_runs, loss_runs = _runs(profits)
    return [
        len(wins),
        len(losses),
        0,
        round(sum(wins), 2),
        round(sum(losses), 2),
        round(_avg(wins), 2),
        round(_avg(losses), 2),
        round(max(wins), 2) if wins else 0,
        round(min(losses), 2) if losses else 0,
        max(win_runs) if win_runs else 0,
        max(loss_runs) if loss_runs else 0,
        round(_avg(win_runs), 2),
        round(_avg(loss_runs), 2),
        0,
        0,
    ]


def _write_horizontal_table(ws, row: int, headers: list[str], values: list[float | int | str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _thin_border()
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row + 1, col, value)
        cell.border = _thin_border()


def _profit_in_pips(report: StrategyReport, trades: list[Trade]) -> float:
    factor = 100 if report.symbol.endswith("JPY") else 10000
    total = 0.0
    for trade in trades:
        sign = 1 if trade.trade_type.lower() == "buy" else -1
        total += (trade.close_price - trade.open_price) * sign * factor
    return total


def _runs(values: list[float]) -> tuple[list[int], list[int]]:
    win_runs: list[int] = []
    loss_runs: list[int] = []
    current_type: str | None = None
    current_len = 0
    for value in values:
        value_type = "win" if value > 0 else "loss" if value < 0 else None
        if value_type is None:
            continue
        if value_type == current_type:
            current_len += 1
        else:
            if current_type == "win":
                win_runs.append(current_len)
            elif current_type == "loss":
                loss_runs.append(current_len)
            current_type = value_type
            current_len = 1
    if current_type == "win":
        win_runs.append(current_len)
    elif current_type == "loss":
        loss_runs.append(current_len)
    return win_runs, loss_runs


def _avg(values: list[float] | list[int]) -> float:
    return sum(values) / len(values) if values else 0


def _safe_div(a: float, b: float) -> float:
    return round(a / b, 2) if b else 0


def _date(value: datetime | None) -> str:
    return value.strftime("%d.%m.%Y") if value else ""


def _datetime_text(value: datetime | None) -> str:
    return value.strftime("%d.%m.%Y %H:%M:%S") if value else ""


def _set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name)[:31] or "Strategy"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        base = cleaned[: 31 - len(str(suffix)) - 1]
        candidate = f"{base}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def _thin_border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)
