from __future__ import annotations

from datetime import datetime
from pathlib import Path
import math
import re
import statistics

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .mt5_report import MONTHS, StrategyReport, Trade


KPI_HEADERS = [
    "Total profit",
    "Profit in pips",
    "Yearly avg profit",
    "Yearly avg % ret",
    "CAGR",
    "# of trades",
    "Sharpe ratio",
    "Profit factor",
    "Return / DD ratio",
    "Winning %",
    "Drawdown",
    "% drawdown",
    "Daily avg profit",
    "Monthly avg profit",
    "Average trade",
    "Annual% / Max DD%",
    "R expectancy",
    "R exp score",
    "SQN",
    "SQN score",
]


def build_workbook(reports: list[StrategyReport], output_path: Path) -> None:
    wb = Workbook()
    index_ws = wb.active
    index_ws.title = "INDEX"
    _write_index(index_ws, reports)

    used_sheet_names = {"INDEX"}
    for report in reports:
        ws = wb.create_sheet(_safe_sheet_name(report.name, used_sheet_names))
        _write_quant_sheet(ws, report)

    wb.save(output_path)


def _write_index(ws, reports: list[StrategyReport]) -> None:
    ws.append(["Strategy", "Symbol", "Timeframe", "Period start", "Period end", "Initial deposit", *KPI_HEADERS, "Report path"])
    _style_header_row(ws, 1, ws.max_column, fill="1F4E78", color="FFFFFF")
    for report in reports:
        stats = _quant_stats(report)
        ws.append(
            [
                report.name,
                report.symbol,
                report.timeframe,
                report.period_start,
                report.period_end,
                report.initial_deposit,
                *[stats[key] for key in KPI_HEADERS],
                str(report.path),
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _set_widths(ws, [28, 12, 12, 13, 13, 14, *([14] * len(KPI_HEADERS)), 72])


def _write_quant_sheet(ws, report: StrategyReport) -> None:
    stats = _quant_stats(report)

    ws.sheet_view.showGridLines = False
    ws["A1"] = "QUANT ANALYZER REPORT"
    ws["A1"].font = Font(bold=True, size=18, color="1F1F1F")
    ws["A2"] = report.name
    ws["A2"].font = Font(bold=True, size=14, color="1F4E78")
    ws["A3"] = "Source: MT5Report - Strategy Tester Report"
    ws["A4"] = f"Symbol {report.symbol}    Period {report.timeframe or 'unknown'} : {report.period_start} - {report.period_end}"
    ws["A5"] = f"Initial deposit {report.initial_deposit:.1f}"
    ws["A6"] = f"Parameters {report.set_path.name if report.set_path else '-'}"

    _write_kpi_grid(ws, stats, start_row=8)
    _write_monthly(ws, report, start_row=22)
    stats_row = 22 + len(report.monthly) + 4
    _write_stats(ws, report, stats, stats_row)
    chart_row = stats_row + 11
    _insert_images(ws, report, chart_row)
    trades_row = chart_row + 16
    _write_trades(ws, report.trades, trades_row)

    ws.freeze_panes = f"A{trades_row + 2}"
    _set_widths(ws, [13, 12, 12, 12, 12, 13, 12, 12, 12, 12, 12, 12, 12, 12])


def _write_kpi_grid(ws, stats: dict[str, float | str], start_row: int) -> None:
    layout = [
        ("TOTAL PROFIT", _money(stats["Total profit"]), "PROFIT IN PIPS", f"{_whole(stats['Profit in pips'])} PIPS", "YRLY AVG PROFIT", _money(stats["Yearly avg profit"])),
        ("YRLY AVG % RET", _pct(stats["Yearly avg % ret"]), "CAGR", _pct(stats["CAGR"]), "# OF TRADES", _whole(stats["# of trades"])),
        ("SHARPE RATIO", _num(stats["Sharpe ratio"]), "PROFIT FACTOR", _num(stats["Profit factor"]), "RETURN / DD RATIO", _num(stats["Return / DD ratio"])),
        ("WINNING %", _pct(stats["Winning %"]), "DRAWDOWN", _money(stats["Drawdown"]), "% DRAWDOWN", _pct(stats["% drawdown"])),
        ("DAILY AVG PROFIT", _money(stats["Daily avg profit"]), "MTHLY AVG PROFIT", _money(stats["Monthly avg profit"]), "AVERAGE TRADE", _money(stats["Average trade"])),
        ("ANNUAL% / MAX DD%", _num(stats["Annual% / Max DD%"]), "R EXPECTANCY", f"{_num(stats['R expectancy'])} R", "R EXP SCORE", f"{_num(stats['R exp score'])} R"),
        ("SQN", _num(stats["SQN"]), "SQN SCORE", _num(stats["SQN score"]), "", ""),
    ]
    label_fill = PatternFill("solid", fgColor="EAF2F8")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    for row_offset, row in enumerate(layout):
        excel_row = start_row + row_offset * 2
        for group in range(3):
            label = row[group * 2]
            value = row[group * 2 + 1]
            if not label:
                continue
            col = 1 + group * 4
            label_cell = ws.cell(excel_row, col, label)
            value_cell = ws.cell(excel_row + 1, col, value)
            ws.merge_cells(start_row=excel_row, start_column=col, end_row=excel_row, end_column=col + 2)
            ws.merge_cells(start_row=excel_row + 1, start_column=col, end_row=excel_row + 1, end_column=col + 2)
            label_cell.font = Font(bold=True, size=9, color="666666")
            label_cell.fill = label_fill
            label_cell.alignment = Alignment(horizontal="center")
            value_cell.font = Font(bold=True, size=13, color="1F1F1F")
            value_cell.fill = value_fill
            value_cell.alignment = Alignment(horizontal="center")
            for c in range(col, col + 3):
                ws.cell(excel_row, c).border = _thin_border()
                ws.cell(excel_row + 1, c).border = _thin_border()


def _write_monthly(ws, report: StrategyReport, start_row: int) -> None:
    ws.cell(start_row, 1, "MONTHLY PERFORMANCE ($)").font = Font(bold=True, size=12)
    headers = ["Year", *MONTHS, "YTD"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row + 1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    row = start_row + 2
    for year in sorted(report.monthly, reverse=True):
        ws.cell(row, 1, year)
        total = 0.0
        for month in range(1, 13):
            value = round(report.monthly.get(year, {}).get(month, 0.0), 2)
            total += value
            ws.cell(row, month + 1, value)
        ws.cell(row, 14, round(total, 2))
        for col in range(1, 15):
            cell = ws.cell(row, col)
            cell.border = _thin_border()
            cell.number_format = '#,##0.00;[Red]-#,##0.00;0'
        row += 1


def _write_stats(ws, report: StrategyReport, stats: dict[str, float | str], start_row: int) -> None:
    wins = [trade.profit_loss for trade in report.trades if trade.profit_loss > 0]
    losses = [trade.profit_loss for trade in report.trades if trade.profit_loss < 0]

    ws.cell(start_row, 1, "STATS").font = Font(bold=True, size=12)
    rows = [
        ("Strategy", "", "", ""),
        ("Wins/Losses Ratio", _safe_div(len(wins), len(losses)), "Payout Ratio (Avg Win/Loss)", _safe_div(_avg(wins), abs(_avg(losses)))),
        ("AHPR", _ahpr(report), "Z-Score", _strip_parens(report.metrics.get("Z-Score", ""))),
        ("Expectancy", stats["Average trade"], "Deviation $", _trade_deviation(report.trades)),
        ("Stagnation in Days", "", "Stagnation in %", ""),
        ("Trades", "", "", ""),
        ("# of Wins", len(wins), "# of Losses", len(losses)),
        ("Gross Profit", _money(sum(wins)), "Gross Loss", _money(sum(losses))),
        ("Average Win", _money(_avg(wins)), "Average Loss", _money(_avg(losses))),
        ("Largest Win", _money(max(wins) if wins else 0), "Largest Loss", _money(min(losses) if losses else 0)),
    ]
    for row_offset, row in enumerate(rows, start=1):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(start_row + row_offset, col, value)
            if col in {1, 3}:
                cell.font = Font(bold=True)
            cell.border = _thin_border()


def _insert_images(ws, report: StrategyReport, start_row: int) -> None:
    ws.cell(start_row, 1, "CHARTS").font = Font(bold=True, size=12)
    anchors = ["A", "F", "K"]
    for idx, (label, path) in enumerate(report.image_paths.items()):
        col = anchors[idx % len(anchors)]
        row = start_row + 1 + (idx // len(anchors)) * 12
        ws[f"{col}{row}"] = label
        ws[f"{col}{row}"].font = Font(bold=True)
        try:
            img = Image(str(path))
            img.width = min(img.width, 330)
            img.height = min(img.height, 120)
            ws.add_image(img, f"{col}{row + 1}")
        except Exception:
            ws[f"{col}{row + 1}"] = str(path)


def _write_trades(ws, trades: list[Trade], start_row: int) -> None:
    ws.cell(start_row, 1, "TRADES").font = Font(bold=True, size=12)
    headers = ["Ticket", "Type", "Open time", "Open price", "Size", "Close time", "Close price", "Profit/Loss", "Comment"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row + 1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.border = _thin_border()
    for row_idx, trade in enumerate(trades, start=start_row + 2):
        values = [
            trade.ticket,
            trade.trade_type,
            trade.open_time.strftime("%d.%m.%Y %H:%M:%S"),
            trade.open_price,
            trade.size,
            trade.close_time.strftime("%d.%m.%Y %H:%M:%S"),
            trade.close_price,
            trade.profit_loss,
            trade.comment,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.border = _thin_border()
            if col == 8:
                cell.number_format = '$ #,##0.00;[Red]$ -#,##0.00;$ 0.00'


def _quant_stats(report: StrategyReport) -> dict[str, float | str]:
    profits = [trade.profit_loss for trade in report.trades]
    wins = [profit for profit in profits if profit > 0]
    losses = [profit for profit in profits if profit < 0]
    total_profit = round(sum(profits), 2)
    gross_profit = sum(wins)
    gross_loss = abs(sum(losses))
    drawdown = _max_balance_drawdown_amount(report)
    drawdown_pct = _relative_drawdown_pct(report)
    years = _years_between(report.period_start, report.period_end)
    months = years * 12 if years else 0.0
    days = years * 365.25 if years else 0.0
    avg_trade = _avg(profits)
    deviation = _trade_deviation(report.trades)
    yearly_avg_profit = total_profit / years if years else 0.0
    yearly_pct = yearly_avg_profit / report.initial_deposit * 100 if report.initial_deposit else 0.0
    cagr = ((report.initial_deposit + total_profit) / report.initial_deposit) ** (1 / years) - 1 if report.initial_deposit and years else 0.0
    sqn = math.sqrt(len(profits)) * avg_trade / deviation if deviation and profits else 0.0

    return {
        "Total profit": round(total_profit, 2),
        "Profit in pips": round(_profit_in_pips(report), 0),
        "Yearly avg profit": round(yearly_avg_profit, 2),
        "Yearly avg % ret": round(yearly_pct, 2),
        "CAGR": round(cagr * 100, 2),
        "# of trades": len(profits),
        "Sharpe ratio": round(avg_trade / deviation, 2) if deviation else 0.0,
        "Profit factor": round(gross_profit / gross_loss, 2) if gross_loss else 0.0,
        "Return / DD ratio": round(total_profit / drawdown, 2) if drawdown else 0.0,
        "Winning %": round(len(wins) / len(profits) * 100, 2) if profits else 0.0,
        "Drawdown": round(drawdown, 2),
        "% drawdown": round(drawdown_pct, 2),
        "Daily avg profit": round(total_profit / days, 2) if days else 0.0,
        "Monthly avg profit": round(total_profit / months, 2) if months else 0.0,
        "Average trade": round(avg_trade, 2),
        "Annual% / Max DD%": round((cagr * 100) / drawdown_pct, 2) if drawdown_pct else 0.0,
        "R expectancy": round(avg_trade / abs(_avg(losses)), 2) if losses else 0.0,
        "R exp score": round((avg_trade / abs(_avg(losses))) * len(profits) / 6.3, 2) if losses else 0.0,
        "SQN": round(sqn, 2),
        "SQN score": round(sqn / 4.4, 2),
    }


def _profit_in_pips(report: StrategyReport) -> float:
    factor = 100 if report.symbol.endswith("JPY") else 10000
    total = 0.0
    for trade in report.trades:
        sign = 1 if trade.trade_type.lower() == "buy" else -1
        total += (trade.close_price - trade.open_price) * sign * factor
    return total


def _max_balance_drawdown_amount(report: StrategyReport) -> float:
    value = _first_metric(report, "Balance Drawdown Maximal", "Reducción máxima del balance")
    amount, _ = _extract_drawdown(value)
    return amount


def _relative_drawdown_pct(report: StrategyReport) -> float:
    value = _first_metric(report, "Balance Drawdown Relative", "Reducción relativa del balance")
    match = re.search(r"([-+]?\d+(?:[.,]\d+)?)%", value)
    if match:
        return _to_float(match.group(1))
    _, pct = _extract_drawdown(_first_metric(report, "Balance Drawdown Maximal", "Reducción máxima del balance"))
    return pct


def _first_metric(report: StrategyReport, *keys: str) -> str:
    for key in keys:
        value = report.metrics.get(key)
        if value:
            return value
    return ""


def _extract_drawdown(value: str) -> tuple[float, float]:
    match = re.search(r"([-+]?\d+(?:[ .]\d{3})*(?:[.,]\d+)?)\s*\(([-+]?\d+(?:[.,]\d+)?)%", value)
    if not match:
        return _to_float(value), 0.0
    return _to_float(match.group(1)), _to_float(match.group(2))


def _trade_deviation(trades: list[Trade]) -> float:
    profits = [trade.profit_loss for trade in trades]
    return statistics.pstdev(profits) if len(profits) > 1 else 0.0


def _years_between(start: str, end: str) -> float:
    try:
        start_dt = datetime.strptime(start, "%d.%m.%Y")
        end_dt = datetime.strptime(end, "%d.%m.%Y")
    except ValueError:
        return 0.0
    return max((end_dt - start_dt).days / 365.25, 0.0)


def _ahpr(report: StrategyReport) -> str:
    value = report.metrics.get("AHPR", "")
    match = re.search(r"\(([-+]?\d+(?:[.,]\d+)?)%\)", value)
    return _pct(_to_float(match.group(1))) if match else value


def _strip_parens(value: str) -> str:
    return value.split(" ", 1)[0] if value else ""


def _avg(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _safe_div(a: float, b: float) -> float:
    return round(a / b, 2) if b else 0.0


def _to_float(value: object) -> float:
    cleaned = str(value).replace(" ", "").replace("%", "").replace(",", ".").strip()
    if not cleaned:
        return 0.0
    match = re.match(r"([-+]?\d+(?:\.\d+)?)", cleaned)
    return float(match.group(1)) if match else 0.0


def _money(value: object) -> str:
    return f"$ {_to_float(value):,.2f}"


def _pct(value: object) -> str:
    return f"{_to_float(value):,.2f} %"


def _num(value: object) -> str:
    return f"{_to_float(value):,.2f}".rstrip("0").rstrip(".")


def _whole(value: object) -> str:
    return f"{_to_float(value):,.0f}".replace(",", "")


def _style_header_row(ws, row_idx: int, max_col: int, fill: str, color: str) -> None:
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(bold=True, color=color)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name)[:31] or "Strategy"
    candidate = cleaned
    suffix = 2
    while candidate in used:
        base = cleaned[: 31 - len(str(suffix)) - 1]
        candidate = f"{base}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def _thin_border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)
